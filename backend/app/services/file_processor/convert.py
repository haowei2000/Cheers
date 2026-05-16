"""File parsing support for txt / md / html / docx / pdf inference flows."""
from __future__ import annotations

import io
from dataclasses import dataclass
from pathlib import Path

from bs4 import BeautifulSoup

PARSABLE_DOCUMENT_TYPES: dict[str, set[str]] = {
    ".txt": {"text/plain", "application/octet-stream"},
    ".md": {"text/markdown", "text/plain", "application/octet-stream"},
    ".html": {"text/html", "text/plain", "application/octet-stream"},
    ".htm": {"text/html", "text/plain", "application/octet-stream"},
    ".docx": {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/octet-stream",
    },
    ".pdf": {"application/pdf", "application/octet-stream"},
    ".xlsx": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
    },
}

KKFILEVIEW_DOCUMENT_TYPES: dict[str, set[str]] = {
    ".doc": {"application/msword", "application/octet-stream"},
    ".xls": {"application/vnd.ms-excel", "application/octet-stream"},
    ".ppt": {"application/vnd.ms-powerpoint", "application/octet-stream"},
    ".pptx": {
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        "application/octet-stream",
    },
    ".wps": {"application/vnd.ms-works", "application/octet-stream"},
    ".et": {"application/octet-stream"},
    ".dps": {"application/octet-stream"},
    ".ofd": {"application/ofd", "application/octet-stream"},
    ".rtf": {"application/rtf", "text/rtf", "application/octet-stream"},
    ".csv": {"text/csv", "text/plain", "application/vnd.ms-excel", "application/octet-stream"},
    ".zip": {"application/zip", "application/x-zip-compressed", "application/octet-stream"},
    ".rar": {"application/vnd.rar", "application/x-rar-compressed", "application/octet-stream"},
    ".7z": {"application/x-7z-compressed", "application/octet-stream"},
    ".tar": {"application/x-tar", "application/octet-stream"},
    ".gz": {"application/gzip", "application/x-gzip", "application/octet-stream"},
    ".bz2": {"application/x-bzip2", "application/octet-stream"},
    ".xz": {"application/x-xz", "application/octet-stream"},
    ".dwg": {"application/acad", "application/x-acad", "application/octet-stream"},
    ".dxf": {"image/vnd.dxf", "application/dxf", "application/octet-stream"},
    ".epub": {"application/epub+zip", "application/octet-stream"},
}

SUPPORTED_DOCUMENT_TYPES: dict[str, set[str]] = {
    **PARSABLE_DOCUMENT_TYPES,
    **KKFILEVIEW_DOCUMENT_TYPES,
}

SUPPORTED_IMAGE_TYPES: dict[str, set[str]] = {
    ".png":  {"image/png"},
    ".jpg":  {"image/jpeg"},
    ".jpeg": {"image/jpeg"},
    ".webp": {"image/webp"},
    ".gif":  {"image/gif"},
}

# Combined supported types for presign validation.
ALL_SUPPORTED_TYPES: dict[str, set[str]] = {**SUPPORTED_DOCUMENT_TYPES, **SUPPORTED_IMAGE_TYPES}


def is_image_type(content_type_or_suffix: str) -> bool:
    """Return whether a MIME type or file extension is an image type."""
    s = content_type_or_suffix.lower()
    return s in SUPPORTED_IMAGE_TYPES or s in {
        mime for mimes in SUPPORTED_IMAGE_TYPES.values() for mime in mimes
    }


class FileParseError(Exception):
    """File parsing failed."""


class UnsupportedFileTypeError(FileParseError):
    """File type is unsupported."""


@dataclass(frozen=True)
class ParsedDocument:
    """Unified file parsing result."""

    text: str
    summary: str
    truncated: bool
    detected_type: str
    char_count: int


def parse_document_bytes(
    payload: bytes,
    *,
    filename: str,
    content_type: str | None = None,
    max_chars: int = 12000,
) -> ParsedDocument:
    """Parse an uploaded file into plain text and truncate it when needed."""

    suffix = Path(filename).suffix.lower()
    if suffix not in PARSABLE_DOCUMENT_TYPES:
        raise UnsupportedFileTypeError(f"unsupported file type: {suffix or '(none)'}")

    normalized_type = (content_type or "").split(";", 1)[0].strip().lower()
    allowed_types = PARSABLE_DOCUMENT_TYPES[suffix]
    if normalized_type and normalized_type not in allowed_types:
        raise UnsupportedFileTypeError(
            f"unsupported content type for {suffix}: {normalized_type}"
        )

    if not payload:
        raise FileParseError("empty file")

    if suffix in (".txt", ".md"):
        text = _parse_text(payload)
    elif suffix in (".html", ".htm"):
        text = _parse_html(payload)
    elif suffix == ".docx":
        text = _parse_docx(payload)
    elif suffix == ".pdf":
        text = _parse_pdf(payload)
    elif suffix == ".xlsx":
        text = _parse_xlsx(payload)
    else:
        raise UnsupportedFileTypeError(f"unsupported file type: {suffix}")

    cleaned = _normalize_text(text)
    if not cleaned:
        raise FileParseError("parsed content is empty")

    truncated_text, truncated = _truncate_text(cleaned, max_chars=max_chars)
    return ParsedDocument(
        text=truncated_text,
        summary=_build_summary(truncated_text),
        truncated=truncated,
        detected_type=suffix.lstrip("."),
        char_count=len(cleaned),
    )


def to_markdown(file_path: str | Path, *, max_chars: int = 12000) -> str:
    """Compatibility API: parse a local path into plain text."""

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(str(path))
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md"}:
        return _normalize_text(path.read_text(encoding="utf-8", errors="replace"))
    if suffix in {".html", ".htm"}:
        return _normalize_text(_parse_html(path.read_bytes()))
    if suffix == ".docx":
        return _normalize_text(_parse_docx(path.read_bytes()))
    if suffix == ".pdf":
        return _normalize_text(_parse_pdf(path.read_bytes()))
    if suffix == ".xlsx":
        return _parse_xlsx(path)
    if suffix in {".png", ".jpg", ".jpeg", ".webp"}:
        return _image_placeholder()
    parsed = parse_document_bytes(
        path.read_bytes(),
        filename=path.name,
        max_chars=max_chars,
    )
    return parsed.text


def _parse_text(payload: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-16", "gb18030"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    return payload.decode("utf-8", errors="replace")


def _parse_html(payload: bytes) -> str:
    soup = BeautifulSoup(_parse_text(payload), "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    return soup.get_text("\n")


def _parse_docx(payload: bytes) -> str:
    import mammoth

    try:
        result = mammoth.extract_raw_text(io.BytesIO(payload))
    except Exception as exc:
        raise FileParseError(f"failed to parse docx: {exc}") from exc
    return result.value


def _parse_pdf(payload: bytes) -> str:
    import pymupdf

    try:
        doc = pymupdf.open(stream=payload, filetype="pdf")
    except Exception as exc:
        raise FileParseError(f"failed to open pdf: {exc}") from exc

    try:
        parts = [page.get_text() for page in doc]
    except Exception as exc:
        raise FileParseError(f"failed to extract pdf text: {exc}") from exc
    finally:
        doc.close()
    return "\n\n".join(parts)


def _normalize_text(text: str) -> str:
    normalized = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    lines = [line.rstrip() for line in normalized.split("\n")]
    return "\n".join(lines).strip()


def _parse_xlsx(source: bytes | Path) -> str:
    from openpyxl import load_workbook

    workbook_source = io.BytesIO(source) if isinstance(source, bytes) else source
    workbook = load_workbook(workbook_source, read_only=True, data_only=True)
    try:
        parts: list[str] = []
        for sheet in workbook.worksheets:
            parts.append(f"## {sheet.title}\n")
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            header = "| " + " | ".join(str(cell) if cell is not None else "" for cell in rows[0]) + " |"
            separator = "| " + " | ".join("---" for _ in rows[0]) + " |"
            parts.append(header)
            parts.append(separator)
            for row in rows[1:21]:
                parts.append("| " + " | ".join(str(cell) if cell is not None else "" for cell in row) + " |")
            if len(rows) > 21:
                parts.append(f"\n*(仅展示前 21 行，共 {len(rows)} 行)*")
            parts.append("")
        return "\n".join(parts).strip() or "(空表)"
    finally:
        workbook.close()


def _image_placeholder() -> str:
    return "[图片内容描述需要接入 Vision / OCR 后生成]"


def _truncate_text(text: str, *, max_chars: int) -> tuple[str, bool]:
    if max_chars <= 0 or len(text) <= max_chars:
        return text, False
    clipped = text[: max_chars - 1].rstrip()
    return clipped + "\n\n[内容已截断，已省略后续文本]", True


def _build_summary(text: str) -> str:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        return ""
    picked = lines[:3]
    return "\n".join(f"- {line[:120]}" for line in picked)
